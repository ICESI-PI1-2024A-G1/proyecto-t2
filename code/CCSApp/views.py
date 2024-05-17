from django.forms import ValidationError
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from .forms import *
from .models import *
import csv
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from django.db.models import Q
from django.views.generic.base import TemplateView
import openpyxl
from django.http import HttpResponse
from django.contrib import messages

# Create your views here.

def asignar_horario(request):
    if request.method == 'POST':
        form = NewHorario(request.POST)
        if form.is_valid():
            fecha_inicio_horario = form.cleaned_data['fecha_inicio_horario']
            hora_inicio_horario = form.cleaned_data['hora_inicio_horario']
            hora_final_horario = form.cleaned_data['hora_final_horario']
            modalidad = form.cleaned_data['modalidad']
            enlace_virtual = form.cleaned_data['enlace_virtual']
            salon_presencial = form.cleaned_data['salon_presencial']
            materia = form.cleaned_data['materia']
            grupo = form.cleaned_data['grupo']

            # Consulta para verificar conflictos
            conflicto = Horario.objects.filter(
                salon_presencial=salon_presencial,
                fecha_inicio_horario=fecha_inicio_horario,  # Coincide en la misma fecha
            ).exclude(
                Q(hora_final_horario=hora_inicio_horario) |  # No se superpone antes
                Q(hora_inicio_horario=hora_final_horario)    # No se superpone después
            ).exists()

            if form.cleaned_data['modalidad'] == 'Virtual':
                # Obtener el edificio y espacio 'N/A'
                nuevo_espacio = Espacio.objects.get(espacio_codigo='N/A')
                # Asociar el nuevo espacio al horario
                salon_presencial = nuevo_espacio

            if not conflicto:
                # No hay conflicto, guardar el nuevo horario
                Horario.objects.create(
                    fecha_inicio_horario=fecha_inicio_horario,
                    hora_inicio_horario=hora_inicio_horario,
                    hora_final_horario=hora_final_horario,
                    materia=materia,
                    modalidad=modalidad,
                    enlace_virtual=enlace_virtual,
                    salon_presencial=salon_presencial,
                    grupo=grupo
                )
                return redirect('/index/programacion')
            else:
                return render(request, 'Asignar/asignar_horario.html', {
                    'formNewHorario': form,
                    'error': 'El horario se superpone con otro horario existente.'
                })

                # Hay conflicto, mostrar un mensaje de error
                #form.add_error(None, "El horario se superpone con otro horario existente.")
    else:
        form = NewHorario()

    return render(request, 'Asignar/asignar_horario.html', {'formNewHorario': form})

def modificar_horarios(request, id_horario):
    horario = Horario.objects.get(pk=id_horario)

    if request.method == 'POST':
        form = ModificarHorarioForm(request.POST, instance=horario)
        if form.is_valid():
            salon_presencial = form.cleaned_data['salon_presencial']
            fecha_inicio_horario = form.cleaned_data['fecha_inicio_horario']
            hora_inicio_horario = form.cleaned_data['hora_inicio_horario']
            hora_final_horario = form.cleaned_data['hora_final_horario']

            if form.cleaned_data['modalidad'] == 'Virtual':
                nuevo_espacio = Espacio.objects.get(espacio_codigo='N/A')
                horario.salon_presencial = nuevo_espacio

            elif form.cleaned_data['modalidad'] == 'Presencial':
                horario.enlace_virtual = 'N/A'

            conflicto = Horario.objects.filter(
                salon_presencial=salon_presencial,
                fecha_inicio_horario=fecha_inicio_horario,
            ).exclude(
                Q(hora_final_horario__lte=hora_inicio_horario) |
                Q(hora_inicio_horario__gte=hora_final_horario) |
                Q(pk=id_horario)
            ).exists()

            if not conflicto:
                form.save()
                return redirect('/index/servicios_asignacion/consultar_horarios')
            else:
                # Si hay conflicto, volver a renderizar el mismo formulario con el mensaje de error
                return render(request, 'Editar/modificar_horarios.html', {
                    'formModificarHorarios': form,
                    'error': 'El horario se superpone con otro horario existente.'
                })
    else:
        form = ModificarHorarioForm(instance=horario)
    return render(request, 'Editar/modificar_horarios.html', {'formModificarHorarios': form})

def consultar_horarios(request):
    # Obtener todos los horarios inicialmente
    horarios = Horario.objects.all()

    # Obtener parámetros de filtro del formulario si están presentes en la solicitud GET
    modalidad = request.GET.get('modalidad')
    materia = request.GET.get('materia')
    grupo = request.GET.get('grupo')
    fecha_inicio_horario = request.GET.get('fecha_inicio_horario')
    hora_inicio_horario = request.GET.get('hora_inicio_horario')
    hora_final_horario = request.GET.get('hora_final_horario')

    # Aplicar filtros según los parámetros proporcionados en el formulario
    if modalidad:
        horarios = horarios.filter(modalidad=modalidad)
    if materia:
        horarios = horarios.filter(materia=materia)
    if grupo:
        horarios = horarios.filter(grupo=grupo)
    if fecha_inicio_horario:
        horarios = horarios.filter(fecha_inicio_horario = fecha_inicio_horario)
    if hora_inicio_horario:
        horarios = horarios.filter(hora_inicio_horario = hora_inicio_horario)
    if hora_final_horario:
        horarios = horarios.filter(hora_final_horario = hora_final_horario)

    # Pasar los horarios filtrados al contexto para mostrar en la plantilla
    context = {
        'horarios': horarios
    }

    return render(request, 'Buscar/consultar_horarios.html', context)

def servicios_asignacion(request):
    if request.method == "GET":
        return render(request, 'Semestre/servicios_asignacion.html')
    

def registrar_materia_malla(request):
    if request.method == "GET":
        return render(request, 'Registrar/registro_materia.html', {
            'form': CrearMateria
        })
    else:
        try:
            form = CrearMateria(request.POST, request.FILES)
            if form.is_valid():
                codigo_materia = form.cleaned_data['codigo_materia']
                nombre_materia = form.cleaned_data['nombre_materia']
                
                # Verificar si ya existe una materia con el mismo código
                if Materia.objects.filter(codigo_materia=codigo_materia).exists():
                    # Si existe, puedes manejar la situación como desees,
                    # por ejemplo, mostrando un mensaje de error
                    return render(request, 'Registrar/registro_materia.html', {
                        'form': CrearMateria,
                        'error': 'La materia ya existe en la base de datos.'
                    })
                # Si el código de la materia no existe, verificar si el nombre de la materia también existe
                elif Materia.objects.filter(nombre_materia=nombre_materia).exists():
                    return render(request, 'registro_materia.html', {
                        'form': CrearMateria,
                        'error': 'La materia ya existe en la base de datos.'
                    })
                else:
                    # Si no existe ninguna materia con el mismo código ni con el mismo nombre,
                    # continuar con la verificación del archivo y la creación de la materia

                    syllabus_file = form.cleaned_data['syllabus']
                    if not syllabus_file.name.endswith('.pdf'):
                        raise ValidationError('El archivo debe ser un PDF.')

                    # Si no existe, crear la nueva materia
                    materia = Materia(
                        nombre_materia=nombre_materia,
                        codigo_materia=codigo_materia,
                        departamento = form.cleaned_data['departamento'],
                        semestre = form.cleaned_data['semestre'],
                        creditos_materia=form.cleaned_data['creditos_materia'],
                        syllabus=syllabus_file,
                        programa_de_posgrado_materia = form.cleaned_data['programa_de_posgrado_materia'],
                    )
                    materia.save()
                    return redirect('/index/servicios_asignacion')
        except ValueError:
            return render(request, 'Registrar/registro_materia.html', {
                'form': CrearMateria,
                'error': 'Por favor, proporcione datos válidos.'
            })
    

def filtrar_materias(request):
    programa_id = request.GET.get('programa')
    semestre_id = request.GET.get('semestre')
    
    print(f"Programa ID: {programa_id}, Semestre ID: {semestre_id}") 
    print(f"Tipo de programa ID: {type(programa_id)}, Tipo de Semestre ID: {type(semestre_id)}")  # Depuración

    materias = Materia.objects.filter(
        programa_de_posgrado_materia=programa_id,
        semestre=semestre_id
    ).values('codigo_materia', 'nombre_materia')  # Solo obtenemos datos necesarios

    print(materias.query)  # Imprime la consulta SQL
    print(f"Materias encontradas: {list(materias)}")

    return JsonResponse(list(materias), safe=False)  # Retornamos JSON


def buscar_materia(request):
    if request.method == 'POST':
        form = MateriaSearchForm(request.POST)
        if form.is_valid():
            nombre_materia = form.cleaned_data['nombre_materia']
            materias = Materia.objects.filter(nombre_materia__istartswith=nombre_materia)
            return render(request, 'Buscar/buscar_materia.html', {'form': form, 'materias': materias})
    else:
        form = MateriaSearchForm()
    return render(request, 'Buscar/buscar_materia.html', {'form': form})

def editar_materia(request, nombre_materia):
    materia = Materia.objects.get(nombre_materia=nombre_materia)
    if request.method == 'POST':
        form = MateriaEditForm(request.POST, instance=materia)
        if form.is_valid():
            form.save()
            return redirect('buscar_materia')
    else:
        form = MateriaEditForm(instance=materia)
    return render(request, 'Editar/editar_materia.html', {'form': form})

def malla_curricular(request):
    if request.method == 'POST':
        form = CrearMallaCurricular(request.POST)
        if form.is_valid():
            # Procesar los datos del formulario y guardar el programa académico
            malla_curricular = Malla_curricular(
                nombre=form.cleaned_data['nombre'],
                descripcion=form.cleaned_data['descripcion'],
                requisitos_previos=form.cleaned_data['requisitos_previos'],
                programa_de_posgrado=form.cleaned_data['programa_de_posgrado'],)
            malla_curricular.save()
            return redirect('/index/servicios_asignacion/registroMateria')  # Redirigir a alguna vista después de guardar el formulario
    else:
        form = CrearMallaCurricular()
    return render(request, 'Semestre/np_malla_curricular.html', {'form' : form})

def nuevo_programa(request):
    if request.method == 'POST':
        form = CrearProgramaAcademico(request.POST)
        if form.is_valid():
            # Obtener el código del programa académico del formulario
            codigo_programa = form.cleaned_data['codigo_programa']
            nombre_programa = form.cleaned_data['nombre_programa']
            
            # Verificar si el código del programa académico ya existe en la base de datos
            if Programa_de_posgrado.objects.filter(codigo_programa=codigo_programa).exists():
                # Si el código ya existe, mostrar un mensaje de error
                return render(request, 'Crear/nuevo_programa.html', {
                        'form': CrearProgramaAcademico,
                        'error': 'El programa de posgrado ya existe en la base de datos.'
                    })
            elif Programa_de_posgrado.objects.filter(nombre_programa=nombre_programa).exists():
                    return render(request, 'Crear/nuevo_programa.html', {
                        'form': CrearProgramaAcademico,
                        'error': 'El programa de posgrado ya existe en la base de datos.'
                    })
            else:
                # Si el código no existe, procesar los datos del formulario y guardar el programa académico
                programa_academico = Programa_de_posgrado(
                    nombre_programa=form.cleaned_data['nombre_programa'],
                    codigo_programa=codigo_programa,
                    fecha_inicio_programa=form.cleaned_data['fecha_inicio_programa'],
                    estado_programa=form.cleaned_data['estado_programa'],
                    duracion_programa=form.cleaned_data['duracion_programa'],
                    facultad_programa=form.cleaned_data['facultad_programa'],
                    modalidad_programa=form.cleaned_data['modalidad_programa'],
                    director_programa=form.cleaned_data['director_programa'],
                )
                programa_academico.save()
                return redirect('/index/gestion')  # Redirigir a alguna vista después de guardar el formulario
    else:
        form = CrearProgramaAcademico()
    return render(request, 'Crear/nuevo_programa.html', {'form': form})


def buscar_programa_academico(request):
    if request.method == 'POST':
        form = BuscarProgramaForm(request.POST)
        if form.is_valid():
            codigo_programa = form.cleaned_data['codigo_programa']
            programas = Programa_de_posgrado.objects.filter(codigo_programa__istartswith=codigo_programa)
            return render(request, 'Buscar/buscar_programa_academico.html', {'form': form, 'programas': programas})
    else:
        form = BuscarProgramaForm()
    return render(request, 'Buscar/buscar_programa_academico.html', {'form': form})

def editar_programa(request, codigo_programa):  
    programa = get_object_or_404(Programa_de_posgrado, codigo_programa=codigo_programa)  
    form = EditarProgramaForm(request.POST, instance=programa)
    if form.is_valid():
        form.save()
        return redirect('/index/gestion/buscar_programa_academico')  
    else:
        form = EditarProgramaForm(instance=programa)
    return render(request, 'Editar/editar_programa.html', {'form': form})

def consultar_programas_academicos(request):
    # Obtener todos los programas académicos con estado "Inactivo"
    programas = Programa_de_posgrado.objects.filter(estado_programa='Inactivo')

    # Pasar los programas académicos filtrados al contexto para mostrar en la plantilla
    context = {
        'programas': programas
    }

    return render(request, 'Buscar/consultar_programas_academicos.html', context)

def log_in(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data['cedula']
            clave = form.cleaned_data['password']
            user = authenticate(request, cedula=cedula, password=clave)
            try:
                usuario = Usuario.objects.get(cedula=cedula, password=clave)
                # Autenticación exitosa, puedes redirigir a una página de inicio o hacer cualquier otra cosa que necesites.
                # Por ejemplo:
                # return redirect('inicio')  # Cambia 'inicio' con el nombre de tu URL de inicio.
                return redirect(index)
            except Usuario.DoesNotExist:
                # Si no se encuentra el usuario, puedes mostrar un mensaje de error o redirigir de nuevo al formulario de inicio de sesión.
                #form.add_error(None, 'Usuario o clave incorrecta, intente de nuevo')
                return render(request, 'login/log_in.html', {
                        'form': LoginForm,
                        'error': 'Usuario o clave incorrecta, intente de nuevo'
                })
    else:
        form = LoginForm()
    return render(request, 'login/log_in.html', {'form': form})

def register_us(request):
    if request.method == 'POST':
        form = NewUsuary(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre']
            cedula = form.cleaned_data['cedula']
            rol = form.cleaned_data['rol']
            departamento = form.cleaned_data['departamento']
            correo_electronico = form.cleaned_data['correo_electronico']
            telefono = form.cleaned_data['telefono']
            password = form.cleaned_data['password']
            # hashed_password = make_password(password)
            
            # Verificar si ya existe un usuario con la cédula proporcionada
            if Usuario.objects.filter(cedula=cedula).exists():
                # Manejar el caso donde ya existe un usuario con la cédula proporcionada
                # Por ejemplo, puedes agregar un mensaje de error al formulario.
                return render(request, 'Registar/register_us.html', {
                        'form': NewUsuary,
                        'error': 'Ya existe un usuario con esta cédula.'
                })
            else:
                # Crear un nuevo usuario con los datos proporcionados
                Usuario.objects.create(
                nombre=nombre,
                cedula=cedula,
                rol=rol,
                departamento=departamento,
                correo_electronico=correo_electronico,
                telefono=telefono,
                password=password
            )
            # Redirigir a la página principal después del registro exitoso
            return redirect('/')
        # Resto del código para manejar el caso cuando el formulario no es válido
        # ...
    else:
        form = NewUsuary()
    return render(request, 'Registrar/register_us.html', {'form': form})

def gestion(request):
    return render(request, 'Semestre/gestion.html')

def index(request):
    if request.method == "GET":
        return render(request, 'login/index.html')
 
def home(request):
    return render(request, 'login/home.html')

def asignar_espacios(request):
    return render(request, 'Asignar/asignar_espacios.html')

def registrar_profesor(request):
    if request.method == "POST":
        form = RegistrarProfesor(request.POST)
        if form.is_valid():
            nombre_profesor = form.cleaned_data['nombre_profesor']
            cedula_profesor = form.cleaned_data['cedula_profesor']
            especializacion_profesor = form.cleaned_data['especializacion_profesor']
            correo_electronico = form.cleaned_data['correo_electronico']
            telefono = form.cleaned_data['telefono']
            
            if Profesor.objects.filter(cedula_profesor=cedula_profesor).exists():
                return render(request, 'Registrar/registro_profesores.html', {
                    'form': form,
                    'error': 'El profesor ya existe en la base de datos.'
                })
            else:
                profesor = Profesor(
                    nombre_profesor=nombre_profesor,
                    cedula_profesor=cedula_profesor,
                    especializacion_profesor=especializacion_profesor,
                    correo_electronico=correo_electronico,
                    telefono=telefono,
                )
                profesor.save()
                return redirect('/index/servicios_asignacion')
    else:
        form = RegistrarProfesor()
    
    return render(request, 'Registrar/registro_profesores.html', {
        'form': form
    })

def buscar_profesor(request):
    if request.method == 'POST':
        form = ProfesorSearchForm(request.POST)
        if form.is_valid():
            nombre_profesor = form.cleaned_data['nombre_profesor']
            profesores = Profesor.objects.filter(nombre_profesor__icontains=nombre_profesor)
            return render(request, 'Buscar/buscar_profesor.html', {'form': form, 'profesores': profesores})
    else:
        form = ProfesorSearchForm()
    return render(request, 'Buscar/buscar_profesor.html', {'form': form})

def editar_profesor(request, nombre_profesor):
    profesor = Profesor.objects.get(nombre_profesor=nombre_profesor)
    if request.method == 'POST':
        form = ProfesorEditForm(request.POST, instance=profesor)
        if form.is_valid():
            form.save()
            return redirect('buscar_profesor')
    else:
        form = ProfesorEditForm(instance=profesor)
    return render(request, 'Editar/editar_profesor.html', {'form': form})

def lista_programas(request):
    programas = Programa_de_posgrado.objects.all()
    directores = Director_de_programa.objects.all()
    return render(request, 'Listas/lista_programas.html', {'programas': programas, 'directores': directores})


def director_programa(request):
    if request.method == 'POST':
        form = DirectorDePrograma(request.POST, request.FILES)
        if form.is_valid():
            # Procesar los datos del formulario y guardar el programa académico
            director_programa = Director_de_programa(
                nombre_director=form.cleaned_data['nombre_director'],
                numero_director=form.cleaned_data['numero_director'],
                correo_director=form.cleaned_data['correo_director'],
                foto_de_perfil=form.cleaned_data['foto_de_perfil'])

            # Guardar la foto
            with open(director_programa.foto_de_perfil.path, 'wb+') as f:
                for chunk in director_programa.foto_de_perfil.chunks():
                    f.write(chunk)

            director_programa.save()
            return redirect('/gestion/nuevoprograma/mallacurricular')  # Redirigir a alguna vista después de guardar el formulario
        else:
            form = DirectorDePrograma(request.POST)
    else:
        form = DirectorDePrograma()
    return render(request, 'Semestre/director_programa.html', {'form': form})

def operacionexitosanp(request):
    return render(request, 'operacion_exitosa_np.html')

def eliminar_programa_inactivo(request, codigo_programa):
    try:
        # Obtener el programa de posgrado con el código proporcionado
        programa = Programa_de_posgrado.objects.get(codigo_programa=codigo_programa)
        
        # Eliminar el programa si existe
        programa.delete()
        
        # Redirigir a una página de confirmación o mostrar un mensaje de éxito
        return redirect('/index/gestion/eliminar_programa_academico')

    except Programa_de_posgrado.DoesNotExist:
        # Manejar el caso en el que el programa no existe
        return render(request, 'Editar/error_programa_inexistente.html')

def delete_program(request, codigo):
    programa = Programa_de_posgrado.objects.get(pk = codigo)
    programa.delete()
    return redirect('eliminar_programa')

def programs_csv(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=programasdeposgrado.xlsx'

    # Crear un nuevo libro de trabajo
    wb = Workbook()
    # Seleccionar la hoja activa
    ws = wb.active

    # Escribir el encabezado
    headers = ['PROGRAMA', 'COD BANNER', 'DEPT', 'HORAS', 'NUM. CREDITOS', 'PERIODO', 'MATERIA', 'MODALIDAD', 'GRUPO', 'DOCENTE', 'C.C', 'TIPO DE CONTRATO', 'CIUDAD', 'EMAIL', 'TELEFONO', 'FECHA DE CLASE', 'HORARIO', 'ESTADO DE CONTRATO', 'FECHA ELAB. DE CONTRATO', 'No. CONTRATO', 'LISTAS - MOSAICOS', 'ENTREGA DE NOTAS', 'INTU/CANVAS', 'TIQUETES', 'HOTEL', 'VIATICOS']
    ws.append(headers)

    # Iterar sobre los programas y escribir los datos en el archivo Excel
    # programas = Programa_de_posgrado.objects.all()
    # for programa in programas:
    #     row_data = [programa.name, programa.codigo, programa.descripcion, programa.fecha_inicio, programa.fecha_finalizacion, programa.estado, programa.duracion, programa.facultad]
    #     ws.append(row_data)

    # Aplicar color a las celdas de los encabezados
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Obtener el índice de la columna "GRUPO" en la hoja de cálculo
    grupo_index = headers.index('GRUPO') + 1

    # Aplicar color verde a las celdas de la columna "GRUPO"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=grupo_index, max_col=grupo_index):
        for cell in row:
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Agregar bordes a todas las celdas
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border

    # Ajustar el ancho de las columnas al contenido
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Obtener la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Multiplicar por 1.2 para dar un poco de espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el libro de trabajo en el flujo de respuesta
    wb.save(response)
    return response

def programas_csv(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=programasdeposgrado.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    column_headers = [
        'Nombre del Programa',
        'Codigo del programa',
        'Fecha de Inicio',
        'Estado',
        'Duracion (Años)',
        'Facultad',
        'Modalidad',
        'Director del Programa'
    ]

    # Escribir encabezados de columnas
    for col_num, header_title in enumerate(column_headers, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header_title

    # Estilo para el título (celda amarilla)
    title_style = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Aplicar estilo amarillo a los encabezados de columnas
    for col_num in range(1, len(column_headers) + 1):
        worksheet.cell(row=1, column=col_num).fill = title_style

    # Escribir datos de programas
    programas = Programa_de_posgrado.objects.all()
    for row_num, programa in enumerate(programas, start=2):
        worksheet.append([
            programa.nombre_programa,
            programa.codigo_programa,
            programa.fecha_inicio_programa,
            programa.estado_programa,
            programa.duracion_programa,
            programa.facultad_programa.nombre_facultad,
            programa.modalidad_programa,
            programa.director_programa.nombre_director
        ])

    # Ajustar el ancho de las columnas después de escribir los datos
    for col_num in range(1, len(column_headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].auto_size = True

    workbook.save(response)
    return response

def horarios(request, codigo_materia):
    horarios = Horario.objects.filter(materia__codigo=codigo_materia)
    context = {'horarios': horarios}
    return render(request, 'Listas/lista_horarios.html', context)

def elegir_semestre(request):
    semestre = Semestre.objects.all()
    context = { 'semestres': semestre }

    return render(request, "Semestre/elegir_semestre.html", context)


def crear_espacio(request):
    if request.method == 'POST':
        form = EspacioForm(request.POST)
        if form.is_valid():
            # Procesar los datos del formulario y guardar el programa académico
            espacio = Espacio(
                espacio_codigo=form.cleaned_data['espacio_codigo'],
                capacidad_espacio=form.cleaned_data['capacidad_espacio'],
                edificio_espacio=form.cleaned_data['edificio_espacio'],
                disponibilidad_espacio=form.cleaned_data['disponibilidad_espacio'],
                tipo=form.cleaned_data['tipo'])
            espacio.save()
            return redirect('/index/servicios_asignacion')  # Redirigir a alguna vista después de guardar el formulario
    else:
        form = EspacioForm()
    return render(request, 'Crear/crear_espacio.html', {'form': form})

def crear_edificio(request):
    if request.method == 'POST':
        form = CrearEdificio(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/index/servicios_asignacion')  
    else:
        form = CrearEdificio()
    return render(request, 'Crear/crear_edificio.html', {'form': form})

def lista_edificios(request):
    # Filtrar los edificios excluyendo aquellos que tienen nombre 'N/A'
    edificios = Edificio.objects.exclude(nombre_edificio='N/A')
    return render(request, 'Listas/lista_edificios.html', {'edificios': edificios})


def lista_espacios(request, nombre_edificio):
    edificio = get_object_or_404(Edificio, pk=nombre_edificio)
    espacios = Espacio.objects.filter(edificio_espacio=edificio)
    return render(request, 'Listas/lista_espacios.html', {'edificio': edificio, 'espacios': espacios})

def editar_espacio(request, espacio_codigo):
    espacio = get_object_or_404(Espacio, pk=espacio_codigo)
    if request.method == 'POST':
        form = EditarEspacio(request.POST, instance=espacio)
        if form.is_valid():
            form.save()
            return redirect('/index/servicios_asignacion/lista_edificios')  # Redirigir a la lista de edificios después de la edición
    else:
        form = EditarEspacio(instance=espacio)
    return render(request, 'Editar/editar_espacio.html', {'form': form})

def crear_programacion_academica(request):
    if request.method == 'POST':
        form = ProgramacionAcademicaForm(request.POST)  # Maneja datos del POST
        if form.is_valid():
            programacion_academica = ProgramacionAcademicaForm(
                programa_de_posgrado =form.cleaned_data['programa_de_posgrado'],
                semestre =form.cleaned_data['semestre'],
                departamento =form.cleaned_data['departamento'],
                estado_programa = form.cleaned_data['estado_programa'],
                materia =form.cleaned_data['materia'],
                horario =form.cleaned_data['horario'],
                grupo =form.cleaned_data['grupo'],
                profesor = form.cleaned_data['profesor'])
            programacion_academica.save()
            return redirect('/index')  # Redirigir a alguna vista después de guardar el formulario
    else:
        form = ProgramacionAcademicaForm()
    return render(request, 'Semestre/programacion_academica.html', {'form': form})

def crear_evento(request):
    if request.method == 'POST':
        formEvent = EventoForm(request.POST)
        if formEvent.is_valid():
            nombre_evento = formEvent.cleaned_data['nombre_evento']
            # Verificar si ya existe un evento con el mismo nombre
            if Evento.objects.filter(nombre_evento=nombre_evento).exists():
                # Si el evento ya existe, puedes manejar la situación como desees,
                # por ejemplo, mostrando un mensaje de error
                return render(request, 'Crear/crear_evento.html', {'formEvent': formEvent, 'error': 'El evento ya existe.'})
            else:
                # Si el evento no existe, guardar el evento
                evento = Evento(
                    nombre_evento=nombre_evento,
                    fecha_inicio_evento=formEvent.cleaned_data['fecha_inicio_evento'],
                    fecha_finalizacion_evento=formEvent.cleaned_data['fecha_finalizacion_evento'],
                    lugar_evento=formEvent.cleaned_data['lugar_evento'],
                    descripcion_evento=formEvent.cleaned_data['descripcion_evento'],
                    programa_de_posgrado_evento=formEvent.cleaned_data['programa_de_posgrado_evento']
                )
                evento.save()
                return redirect('/index/servicios_asignacion')  # Redirigir a alguna vista después de guardar el formulario
    else:
        formEvent = EventoForm()
    return render(request, 'Crear/crear_evento.html', {'formEvent': formEvent})


def crear_actividad(request):
    if request.method == 'POST':
        formActividad = ActividadForm(request.POST)
        if formActividad.is_valid():
            # Procesar los datos del formulario y guardar el programa académico
            actividad = Actividad(
                nombre_actividad=formActividad.cleaned_data['nombre_actividad'],
                duracion_en_horas=formActividad.cleaned_data['duracion_en_horas'],
                orador_actividad=formActividad.cleaned_data['orador_actividad'],
                evento_actividad=formActividad.cleaned_data['evento_actividad'])
            actividad.save()
            return redirect('/index/servicios_asignacion')  # Redirigir a alguna vista después de guardar el formulario
    else:
        formActividad = ActividadForm()
    return render(request, 'Crear/crear_actividad.html', {'formActividad': formActividad})


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from django.http import HttpResponse

class InformeProgramacion(TemplateView):
    def get(self, request, *args, **kwargs):
        query = ProgramacionAcademica.objects.all()
        wb = Workbook()

        # Crear una nueva hoja en el libro de trabajo
        ws = wb.active
        ws.title = 'Hoja1'

        # Escribir el encabezado
        headers = ['PROGRAMA', 'COD BANNER', 'DEPT', 'HORAS', 'NUM. CREDITOS', 'PERIODO', 'MATERIA', 'MODALIDAD', 'GRUPO', 'DOCENTE', 'C.C', 'TIPO DE CONTRATO', 'CIUDAD', 'EMAIL', 'TELEFONO', 'FECHA DE CLASE', 'HORARIO', 'ESTADO DE CONTRATO', 'FECHA ELAB. DE CONTRATO', 'No. CONTRATO', 'LISTAS - MOSAICOS', 'ENTREGA DE NOTAS', 'INTU/CANVAS', 'TIQUETES', 'HOTEL', 'VIATICOS']
        ws.append(headers)

        # Aplicar formato al encabezado
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Obtener los datos de los registros y escribirlos en el archivo xlsx
        for registro in query:
            data_row = [
                registro.programa_de_posgrado.nombre_programa,
                registro.cod_banner,
                registro.departamento.id_departamento,
                registro.horas,
                registro.num_creditos,
                registro.periodo.id_periodo,
                ', '.join([materia.nombre_materia for materia in registro.materia.all()]),  # Obtener los nombres de las materias separados por coma
                registro.modalidad,
                registro.grupo,
                registro.docente.nombre_profesor,
                registro.tipo_de_contrato,
                registro.ciudad,
                registro.correo_electronico,
                registro.telefono,
                registro.fecha_de_clase,
                ', '.join([str(horario.id_horario) for horario in registro.horario.all()]),  # Obtener los nombres de los horarios separados por coma
                registro.estado_de_contrato,
                registro.fecha_elab_contrato,
                registro.num_contrato,
                registro.listas_mosaicos,
                registro.entrega_notas,
                registro.intu_canvas,
                registro.tiquetes,
                registro.hotel,
                registro.viaticos
            ]
            ws.append(data_row)

        # Aplicar bordes a todas las celdas
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Ajustar el ancho de las columnas al contenido
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Obtener la letra de la columna
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Multiplicar por 1.2 para dar un poco de espacio adicional
            ws.column_dimensions[column_letter].width = adjusted_width

        # Configurar la respuesta HTTP
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename=ReporteProgramacionAcademica.xlsx"
        response["Content-Disposition"] = contenido

        # Guardar el archivo xlsx y devolver la respuesta
        wb.save(response)
        return response





